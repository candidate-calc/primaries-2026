# -*- coding: utf-8 -*-
# מייצר docs/data.js לאתר: דירוגים (מה-CSV) + תוכן קובצי ה-MD.
import csv, json, os, glob, shutil
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOCS = os.path.join(BASE, 'docs')
os.makedirs(DOCS, exist_ok=True)

# ברירות המחדל של האתר = הפרמטרים השמורים בגיליון "פרמטרים" באקסל (מקור האמת)
_pw = load_workbook(os.path.join(BASE, 'primaries-scoring.xlsx'), data_only=True)['פרמטרים']
DEFAULTS = {
    'wT':    [_pw.cell(4+k, 2).value for k in range(10)],
    'wC':    [_pw.cell(16+k, 2).value for k in range(5)],
    'alpha': [_pw.cell(23, 2+k).value for k in range(5)],
    'gamma': [_pw.cell(26, 2+k).value for k in range(4)],
    'lam':   _pw.cell(28, 2).value,
    'beta':  _pw.cell(29, 2).value,
}
assert all(v is not None for v in DEFAULTS['wT'] + DEFAULTS['wC'] + DEFAULTS['alpha'] + DEFAULTS['gamma'] + [DEFAULTS['lam'], DEFAULTS['beta']])

TOPICS = ["ביטחון וחטופים","דמוקרטיה ומשפט","דת ומדינה ושוויון בנטל","הסדרה ושלום",
          "שוויון אזרחי ושותפות","כלכלה עבודה ורווחה","חינוך ובריאות","סביבה ואקלים",
          "שקיפות וטוהר מידות","מגדר ולהט\"ב"]
TRAITS = ["יושרה","חריצות וביצוע","אומץ ציבורי","חיבור ועבודה רוחבית","מקצועיות וניסיון"]
TDIMS = ["הצהרות","תקשורת","שטח","הצלחה","קרדיט"]
CDIMS = ["עדות עצמית","קרדיט עמיתים","תקשורת","שטח"]

def read_csv(name):
    with open(os.path.join(BASE,'research',name), encoding='utf-8-sig') as f:
        return list(csv.reader(f))

trows = read_csv('ratings-topics.csv')[1:]
crows = read_csv('ratings-traits.csv')[1:]

# ערכי ויקיפדיה עברית קיימים (כותרות מאומתות מול ה-API, research/wiki-extracts.txt)
WIKI = {1:'משה רדמן אבוטבול',2:'ערן עציון',3:'מורן זר קצנשטיין',4:'יעל כהן-פארן',6:'רם שפע',
 7:'יריב אופנהיימר',9:'נאור נרקיס',10:'מהרטה ברוך-רון',12:'ירון ניב',14:'אליס גולדמן',
 15:'אבי דבוש',18:'ענבר בזק',22:"אמיר ח'ניפס",23:'הדס רגולסקי',25:'חן אריאלי',26:'מיכל רוזין',
 28:'אפרת רייטן',33:'יאיא פינק',35:'נאווה רוזוליו',36:'מאלכ בדר',38:'גלעד קריב',41:'אמילי מואטי',
 43:'תומר אביטל',44:'סומיה בשיר',46:'גבי לסקי',47:'נעמה לזימי',48:'איהאב שליאן',49:'מוסי רז',
 50:'עלי סלאלחה',51:'נמרוד שפר'}

import re as _re
def personal_site(md):
    m = _re.search(r'אתר אישי: ([^\s·\n]+)', md)
    if not m: return ''
    tok = m.group(1).strip()
    if '.' not in tok or tok.startswith('אין') or 'ויקיפדיה' in tok or tok.startswith('site123'):
        return ''
    return tok if tok.startswith('http') else 'https://' + tok

cands = []
cmap = {int(r[0]): r for r in crows}
dossier_files = {int(os.path.basename(p).split('-')[0]): os.path.basename(p)
                 for p in glob.glob(os.path.join(BASE,'candidates','*.md'))}
for r in trows:
    i = int(r[0]); name = r[1]; ev = float(r[2])
    vals = [float(x) for x in r[3:]]
    T = [vals[t*5:(t+1)*5] for t in range(10)]
    cv = [float(x) for x in cmap[i][2:]]
    C = [cv[c*4:(c+1)*4] for c in range(5)]
    photo = f'photos/{i:02d}.jpg'
    if not os.path.exists(os.path.join(DOCS, 'photos', f'{i:02d}.jpg')): photo = ''
    md_path = os.path.join(BASE,'candidates', dossier_files.get(i,''))
    site = ''
    if dossier_files.get(i):
        with open(md_path, encoding='utf-8') as f: site = personal_site(f.read())
    wiki = 'https://he.wikipedia.org/wiki/' + WIKI[i].replace(' ','_') if i in WIKI else ''
    cands.append({'id': i, 'name': name, 'ev': ev, 'T': T, 'C': C, 'file': dossier_files.get(i,''),
                  'photo': photo, 'wiki': wiki, 'site': site})

docs = {}
for fn in ['README.md','01-plan.md','02-sources.md','03-topics-traits.md','04-scoring-method.md']:
    with open(os.path.join(BASE,fn), encoding='utf-8') as f: docs[fn] = f.read()
dossiers = {}
for i, fn in dossier_files.items():
    with open(os.path.join(BASE,'candidates',fn), encoding='utf-8') as f: dossiers[i] = f.read()

data = {
 'topics': TOPICS, 'traits': TRAITS, 'tdims': TDIMS, 'cdims': CDIMS,
 'candidates': cands,
 'defaults': DEFAULTS,
 'docs': docs, 'dossiers': dossiers,
 'meta': {'collected':'6.7.2026','vote':'20.7.2026'}
}
with open(os.path.join(DOCS,'data.js'),'w',encoding='utf-8') as f:
    f.write('window.PRIMARIES_DATA = ')
    json.dump(data, f, ensure_ascii=False)
    f.write(';')

shutil.copy2(os.path.join(BASE,'primaries-scoring.xlsx'), os.path.join(DOCS,'primaries-scoring.xlsx'))
print('data.js written,', len(cands), 'candidates,', len(dossiers), 'dossiers')
